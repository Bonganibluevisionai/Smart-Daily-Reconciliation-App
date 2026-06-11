from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from extractor import extract_recon_data


SHIFT_COLUMN_MAP = {
    ("Morning", "1"): 2,
    ("Morning", "2"): 3,
    ("Morning", "3"): 4,
    ("Day", "1"): 5,
    ("Day", "2"): 6,
    ("Day", "3"): 7,
    ("Night", "1"): 8,
    ("Night", "2"): 9,
    ("Night", "3"): 10,
}


@dataclass
class WorkbookSlot:
    shift: str
    pos: str
    cashier_name: str
    total_cash: float
    speedpoint_total: float
    safe_drop_value: float
    theoretical_cashier_total: float
    hard_cash_should_till: float
    surplus_shortage: float

    def as_dict(self) -> dict[str, Any]:
        return {
            "shift": self.shift,
            "pos": self.pos,
            "cashier_name": self.cashier_name,
            "total_cash": self.total_cash,
            "speedpoint_total": self.speedpoint_total,
            "safe_drop_value": self.safe_drop_value,
            "theoretical_cashier_total": self.theoretical_cashier_total,
            "hard_cash_should_till": self.hard_cash_should_till,
            "surplus_shortage": self.surplus_shortage,
        }


def _to_float(value: Any) -> float:
    if value in (None, "", "#REF!"):
        return 0.0
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def _normalized_label(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _find_row_index(sheet: Any, label: str) -> int:
    target = _normalized_label(label)
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        first_cell = _normalized_label(row[0] if row else "")
        if first_cell == target:
            return row_index
    raise ValueError(f"Could not find row '{label}' in workbook sheet '{sheet.title}'")


def load_cashup_slots(workbook_path: str | Path) -> list[WorkbookSlot]:
    workbook = load_workbook(filename=str(workbook_path), data_only=True)
    sheet = workbook["Cashier's Cash Up (MoP)"]

    cashier_row = _find_row_index(sheet, "Cashier Names")
    total_cash_row = _find_row_index(sheet, "TOTAL CASH FOR THE SHIFT")
    speedpoint_row = _find_row_index(sheet, "SPEED POINT TOTALS")
    safe_drop_row = _find_row_index(sheet, "SAFE DROP VALUE")
    theoretical_row = _find_row_index(sheet, "THEORATICAL CASHIER TOTAL")
    hard_cash_row = _find_row_index(sheet, "HARD CASH THAT SHOULD BE ON TILL")
    surplus_row = _find_row_index(sheet, "SURPLUS/SHORTAGE")

    slots: list[WorkbookSlot] = []
    for (shift, pos), column_index in SHIFT_COLUMN_MAP.items():
        total_cash = _to_float(sheet.cell(total_cash_row, column_index).value)
        if total_cash == 0.0:
            continue

        slots.append(
            WorkbookSlot(
                shift=shift,
                pos=pos,
                cashier_name=str(sheet.cell(cashier_row, column_index).value or "").strip(),
                total_cash=total_cash,
                speedpoint_total=_to_float(sheet.cell(speedpoint_row, column_index).value),
                safe_drop_value=_to_float(sheet.cell(safe_drop_row, column_index).value),
                theoretical_cashier_total=_to_float(sheet.cell(theoretical_row, column_index).value),
                hard_cash_should_till=_to_float(sheet.cell(hard_cash_row, column_index).value),
                surplus_shortage=_to_float(sheet.cell(surplus_row, column_index).value),
            )
        )

    return slots


def _site_totals(slots: list[WorkbookSlot]) -> dict[str, float]:
    return {
        "total_cash": round(sum(slot.total_cash for slot in slots), 2),
        "speedpoint_total": round(sum(slot.speedpoint_total for slot in slots), 2),
        "safe_drop_value": round(sum(slot.safe_drop_value for slot in slots), 2),
        "theoretical_cashier_total": round(sum(slot.theoretical_cashier_total for slot in slots), 2),
        "hard_cash_should_till": round(sum(slot.hard_cash_should_till for slot in slots), 2),
        "surplus_shortage": round(sum(slot.surplus_shortage for slot in slots), 2),
    }


def _metric_triplet(label: str, pdf_value: float, excel_value: float) -> dict[str, Any]:
    delta = round(pdf_value - excel_value, 2)
    return {
        "label": label,
        "pdf_value": round(pdf_value, 2),
        "excel_value": round(excel_value, 2),
        "delta": delta,
        "matches": abs(delta) < 0.01,
    }


def _match_reports_to_slots(reports: list[dict[str, Any]], slots: list[WorkbookSlot]) -> tuple[list[tuple[dict[str, Any], WorkbookSlot]], list[dict[str, Any]], list[WorkbookSlot]]:
    matched: list[tuple[dict[str, Any], WorkbookSlot]] = []
    unmatched_reports = list(reports)
    available_slots = list(slots)

    for report in list(unmatched_reports):
        report_total = round(float(report.get("cashier_recon", {}).get("Total cash for the Shift", 0.0)), 2)
        candidates = [slot for slot in available_slots if abs(slot.total_cash - report_total) < 0.01]
        if len(candidates) == 1:
            slot = candidates[0]
            matched.append((report, slot))
            unmatched_reports.remove(report)
            available_slots.remove(slot)

    for report in list(unmatched_reports):
        report_shift = str(report.get("meta", {}).get("shift_label") or "").title()
        report_pos = str(report.get("meta", {}).get("pos_terminal") or "").strip()
        slot = next((candidate for candidate in available_slots if candidate.shift == report_shift and candidate.pos == report_pos), None)
        if slot is None and available_slots:
            slot = available_slots[0]
        if slot is not None:
            matched.append((report, slot))
            unmatched_reports.remove(report)
            available_slots.remove(slot)

    return matched, unmatched_reports, available_slots


def build_comparison_report(pdf_paths: list[str | Path], workbook_path: str | Path) -> dict[str, Any]:
    workbook_slots = load_cashup_slots(workbook_path)
    pdf_reports = [extract_recon_data(path) for path in pdf_paths]

    matched, unmatched_reports, unmatched_slots = _match_reports_to_slots(pdf_reports, workbook_slots)

    shift_reports: list[dict[str, Any]] = []
    for report, slot in matched:
        pdf_cashier_recon = report.get("cashier_recon", {})
        pdf_payments = report.get("payments", {})
        pdf_meta = report.get("meta", {})

        metrics = [
            _metric_triplet(
                "Total Cash For Shift",
                float(pdf_cashier_recon.get("Total cash for the Shift", 0.0)),
                slot.total_cash,
            ),
            _metric_triplet(
                "Speedpoint Total",
                float(pdf_cashier_recon.get("Speed point totals") or pdf_cashier_recon.get("Speed ponit totals") or pdf_payments.get("Speedpoint") or 0.0),
                slot.speedpoint_total,
            ),
            _metric_triplet(
                "Safe Drop Value",
                float(pdf_payments.get("Safe Drops", 0.0)),
                slot.safe_drop_value,
            ),
            _metric_triplet(
                "Theoretical Cashier Total",
                float(pdf_cashier_recon.get("Theoritical Cashier Total", 0.0)),
                slot.theoretical_cashier_total,
            ),
        ]

        shift_reports.append(
            {
                "pdf": {
                    "source_file": report.get("source_file", ""),
                    "shift_label": pdf_meta.get("shift_label", "Unknown"),
                    "pos_terminal": str(pdf_meta.get("pos_terminal", "")),
                    "operator": pdf_meta.get("operator", ""),
                },
                "excel": slot.as_dict(),
                "matched_by": "total_cash" if abs(float(pdf_cashier_recon.get("Total cash for the Shift", 0.0)) - slot.total_cash) < 0.01 else "fallback",
                "metrics": metrics,
                "all_match": all(metric["matches"] for metric in metrics),
            }
        )

    pdf_site_totals = {
        "total_cash": round(sum(float(report.get("cashier_recon", {}).get("Total cash for the Shift", 0.0)) for report in pdf_reports), 2),
        "speedpoint_total": round(sum(float(report.get("cashier_recon", {}).get("Speed point totals") or report.get("cashier_recon", {}).get("Speed ponit totals") or report.get("payments", {}).get("Speedpoint") or 0.0) for report in pdf_reports), 2),
        "safe_drop_value": round(sum(float(report.get("payments", {}).get("Safe Drops", 0.0)) for report in pdf_reports), 2),
        "theoretical_cashier_total": round(sum(float(report.get("cashier_recon", {}).get("Theoritical Cashier Total", 0.0)) for report in pdf_reports), 2),
    }
    excel_site_totals = _site_totals(workbook_slots)
    site_metrics = [
        _metric_triplet("Total Cash For Shift", pdf_site_totals["total_cash"], excel_site_totals["total_cash"]),
        _metric_triplet("Speedpoint Total", pdf_site_totals["speedpoint_total"], excel_site_totals["speedpoint_total"]),
        _metric_triplet("Safe Drop Value", pdf_site_totals["safe_drop_value"], excel_site_totals["safe_drop_value"]),
        _metric_triplet("Theoretical Cashier Total", pdf_site_totals["theoretical_cashier_total"], excel_site_totals["theoretical_cashier_total"]),
    ]

    recommendations: list[str] = []
    safe_drop_metric = next((metric for metric in site_metrics if metric["label"] == "Safe Drop Value"), None)
    if safe_drop_metric and not safe_drop_metric["matches"]:
        recommendations.append("Safe drops still differ between PDFs and the workbook. Use the workbook Safe Drop (PDI) values as the final reconciliation source when closing tills.")
    if any(report["pdf"]["shift_label"] != report["excel"]["shift"] for report in shift_reports):
        recommendations.append("At least one PDF shift label does not match the workbook slot. Match reports to workbook rows by total cash before relying on shift labels.")
    if all(metric["matches"] for metric in site_metrics):
        recommendations.append("Workbook and PDF totals align for the uploaded bundle. You can use the PDFs directly for this reconciliation run.")

    return {
        "excel_file": Path(workbook_path).name,
        "pdf_files": [Path(path).name for path in pdf_paths],
        "matched_reports": shift_reports,
        "site_comparison": {
            "pdf_totals": pdf_site_totals,
            "excel_totals": excel_site_totals,
            "metrics": site_metrics,
            "all_match": all(metric["matches"] for metric in site_metrics),
        },
        "unmatched_reports": [report.get("source_file", "") for report in unmatched_reports],
        "unmatched_workbook_slots": [slot.as_dict() for slot in unmatched_slots],
        "recommendations": recommendations,
    }